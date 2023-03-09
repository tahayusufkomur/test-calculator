import json
import os
import pathlib

import numpy as np
import openpyxl.chart.axis
from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.layout import ManualLayout, Layout
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from thefuzz import fuzz

from src.utilities import is_question, get_number_from_question, \
    flat_text_map, convert_question_columns_to_number, get_contradiction, intersection, get_question_with_number, logger
import pandas as pd
from datetime import datetime
import string


class BasePersonalityTest:
    scoring = None
    reverse_scoring = None
    contradiction_score = 0
    round = 0  # round doubles
    date = datetime.today().strftime('%Y-%m-%d-%H-%M')
    copy_df = None
    category_means = {}
    name = ""
    should_be_same = None
    font = '14pt;'
    text_calculation_df = None
    categories_dict = None
    current_path = pathlib.Path(__file__).parent.resolve()
    if not os.path.isdir(f'{current_path}/raporlar'):
        os.mkdir(f'{current_path}/raporlar')

    passwords_path = f"{current_path}/files/passwords.json"

    should_be_same = [
    ]

    text_scoring = {
    }

    reverse_calculation_list = [
    ]

    reverse_scoring = {
    }

    def __init__(self,
                 test_path: str,
                 colors=None,
                 output_path=None,
                 text_output_path=None,
                 text_output_person=None,
                 temp_dir=None,
                 passwords_path=None,
                 ):

        if colors is None:
            colors = {'very_bad': "#FF00F4", 'bad': "#CD7F32", 'okey': "#C9CC3F", 'good': "#4CBB17",
                      'very_good': "#40E0D0"}

        # output and temp working dirs
        self.passwords = json.load(open(passwords_path, encoding="utf-8")) if passwords_path else json.load(open(f"{self.current_path}/files/passwords.json", encoding="utf-8"))
        self.output_path = f"{self.current_path}/raporlar/excel_reports/{self.name}-rapor-{self.date}.xlsx" if not output_path else output_path + f"/{self.name}.xlsx"
        self.text_output_dir = f"{self.current_path}/raporlar/text_reports" if not text_output_path else text_output_path
        self.text_output_person = f"{self.text_output_dir}/person_reports/{self.name}-rapor-{self.date}.docx" if not text_output_person else text_output_person
        self.temp_dir = f"{self.current_path}/raporlar/temp_files" if not temp_dir else temp_dir
        self.test_path = test_path
        self.categories = [v[0] for i, v in self.categories_dict.items()]
        self.categories_t = [v[0] + " T" for i, v in self.categories_dict.items()]
        self.test_df = self.read_test_answers()
        self.personal_questions, self.test_questions = self.parse_test()
        self.result_df = None
        self.EXTENDED_UPPERCASE_ALPHABET = list(string.ascii_uppercase) + ['A' + i for i in string.ascii_uppercase]
        self.colors = colors
        self.text_file_path = f"{self.current_path}/files/text_mapping_jsons/{self.name}.json"
        print(self.text_file_path)

    def read_test_answers(self):
        return pd.read_excel(self.test_path, engine='openpyxl')

    def fix_excel_names(self, path, col):
        names = []
        mapper = {}
        df = pd.read_excel(path, engine='openpyxl')
        for i, v in self.passwords.items():
            names.append(v)

        fuzz_names = set(df[col].values)
        names = set(names)

        for name in names:
            for fuzz_name in fuzz_names:
                if 100 > fuzz.ratio(fuzz_name, name) > 85:
                    mapper[fuzz_name] = name
        df[col] = df[col].apply(lambda x: mapper[x] if x in mapper else x)
        df.to_excel(path, engine='openpyxl', index=False)

    def parse_test(self):
        columns = self.test_df.columns
        personal_questions = []
        test_questions = []

        for column in columns:
            if is_question(column):
                test_questions.append(column)
            else:
                personal_questions.append(column)
        return personal_questions, test_questions

    def filter_df_by_password(self):
        self.test_df = self.test_df[self.test_df['Şifre'].str.contains('|'.join(self.passwords))]
        # standardize name column
        self.test_df['İsim Soyisim'] = self.test_df['Şifre'].apply(lambda x: self.passwords[x] if x in self.passwords else x)

    def map_answers(self):

        for column in self.test_questions:
            self.test_df[column] = self.test_df[column].map(self.scoring)

    def map_df(self, df):
        if df is not None:
            for key, row in self.categories_dict.items():
                for column in self.test_questions:
                    if get_number_from_question(column) in row[1]:
                        if get_number_from_question(column) in self.reverse_calculation_list:
                            df[column] = df[column].map(self.reverse_scoring)
                        else:
                            df[column] = df[column].map(self.scoring)
            return df

    def prepare_result(self):
        self.result_df = self.test_df.copy()[self.personal_questions]
        for key, row in self.categories_dict.items():
            self.result_df[row[0]] = (self.sum_row(row[1]) / len(row[1])) * 10
            self.test_df[row[0]] = (self.sum_row(row[1]) / len(row[1])) * 10
        # will be used for getting highest score categories
        self.text_calculation_df = self.test_df.copy()

    def sum_row(self, column_map):
        sum = 0
        for column in self.test_questions:
            number = get_number_from_question(column)
            if number in column_map:
                sum += self.test_df[column]
        return sum

    def get_contradiction_scores(self):

        df = pd.read_excel(self.test_path, engine='openpyxl')
        df = self.filter_df_by_password_return(df)

        df = self.map_df(df)
        df = convert_question_columns_to_number(df)

        contradictions = []
        max_difference_between_answers = max(self.scoring.values()) - min(self.scoring.values())
        potential_contradiction = len(self.should_be_same * max_difference_between_answers)

        for index, row in df.iterrows():
            contradiction_score = 0
            for check_list in self.should_be_same:
                contradiction_score = get_contradiction(row[check_list], check_list, contradiction_score)[
                    'contradiction_score']
            contradiction_score = round((100 - (contradiction_score / potential_contradiction) * 100), 0)
            contradictions.append(contradiction_score)
        return contradictions

    def build_contradiction_df(self, row_name='contradiction_list'):

        df = pd.read_excel(self.test_path, engine='openpyxl')
        df = self.filter_df_by_password_return(df)

        df = self.map_df(df)
        df = convert_question_columns_to_number(df)
        # convert column to np.array function
        df[row_name] = np.array
        # execute the function with empty array
        df[row_name] = df[row_name].apply(lambda x: (x([])))

        for check_list in self.should_be_same:
            df[row_name] = df.apply(lambda row:
                                    np.append(
                                        row[row_name],
                                        get_contradiction(
                                            row[check_list],
                                            check_list,
                                            0)['cols']
                                    ),
                                    axis=1)

        df[row_name] = df[row_name].apply(lambda cols: [i for i in cols if i != 0])
        categorised_contradictions = {}

        for index, row in df.iterrows():
            temp_dict = {}
            for i, value in self.categories_dict.items():
                temp_dict[value[0]] = []

            contradiction_list = row[row_name]
            for grp in contradiction_list:
                for index, value in self.categories_dict.items():
                    category_name = value[0]
                    if intersection(grp, value[1]):
                        temp_dict[category_name] = list(temp_dict[category_name]) + list(grp)
            categorised_contradictions[row['İsim Soyisim']] = temp_dict

        for i, value in self.categories_dict.items():
            categorised_contradictions[value[0]] = []

        df = df[['İsim Soyisim']]

        for index, row in df.iterrows():
            if row['İsim Soyisim'] in categorised_contradictions.keys():
                categories = categorised_contradictions[row['İsim Soyisim']]
                for category, values in categories.items():
                    df.loc[index, category] = str(values)
        return df

    def concat_row(self, column_map, df):
        text = ""
        for column in self.test_questions:
            if get_number_from_question(column) in column_map:
                text += df[column] + " "
        return text

    def add_mean_row(self):
        self.result_df.loc['ortalama'] = self.result_df.mean(numeric_only=True)

    def add_std_row(self):
        self.result_df.loc['standart_sapma'] = self.result_df.std(numeric_only=True)

    def add_tscore(self):
        for key, row in self.categories_dict.items():
            self.result_df[f"{row[0]} T"] = ((self.test_df[row[0]] - self.test_df[row[0]].mean()) / self.test_df[
                row[0]].std()) * 10 + 50

    def add_trust_score(self):
        self.result_df['Güven Endeksi'] = self.get_contradiction_scores()

    def round_df(self):
        self.result_df = self.result_df.round(self.round)

    def adjust_width(self, path=None):
        from openpyxl import load_workbook
        if path:
            path = path
        else:
            path = self.output_path

        wb = load_workbook(path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")

            excel_keymap = {}
            for i, column in enumerate(df.columns):
                excel_keymap[column] = self.EXTENDED_UPPERCASE_ALPHABET[i]
            # Auto-adjust columns' width
            for column in df.columns:
                column_width = max(df[column].astype(str).map(len).max(), len(column))
                sheet.column_dimensions[excel_keymap[column]].width = column_width - 1
        wb.save(path)

    def format_report(self):
        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)

        for sheet_name in wb.sheetnames:
            df = pd.read_excel(self.output_path, engine='openpyxl', sheet_name=sheet_name)
            writer = pd.ExcelWriter(self.output_path, engine='openpyxl', mode='a')
            df = df.style.applymap(self.adjust_font)
            writer.book = wb
            wb.remove(wb[sheet_name])
            df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
            writer.close()

    def save_report(self):
        self.result_df.to_excel(self.output_path, engine='openpyxl', index=False)

    def create_text_report(self):
        text_map = json.load(open(self.text_file_path, encoding="utf-8"))
        # flat json to 1->1-5 to 1_1-5 for later mapping
        text_map = flat_text_map(text_map)

        self.test_df = self.read_test_answers()
        self.filter_df_by_password()
        self.map_answers()

        # map answers to 1-5
        text_df = self.test_df.copy()
        original_df = pd.read_excel(self.test_path, engine='openpyxl')
        original_df = self.filter_df_by_password_return(original_df)
        original_df = convert_question_columns_to_number(original_df)

        # only name and questions
        text_df = text_df[['İsim Soyisim'] + self.test_questions]

        text_df[self.test_questions] = text_df[self.test_questions].stack().map(self.text_scoring).unstack()

        # map answers to q_a ex: q:11, a:5 -> 11_5
        for question in self.test_questions:
            number = get_number_from_question(question)
            text_df[question] = text_df[question].apply(lambda a: f"{number}_{a}")
        # q_a -> text
        text_df[self.test_questions] = text_df[self.test_questions].stack().map(text_map).unstack()

        # drop na columns
        text_df.fillna('', inplace=True)
        text_df_not_concatted = text_df.copy()
        # text -> concat of category
        for key, row in self.categories_dict.items():
            text_df[row[0]] = self.concat_row(row[1], text_df)

        text_df.drop(self.test_questions, axis=1, inplace=True)

        # for latest text report for person
        text_df.to_excel(f"{self.temp_dir}/{self.name}-text.xlsx", engine="openpyxl", index=False)
        categories = [v[0] for v in self.categories_dict.values()]
        contradiction_df = self.build_contradiction_df(self.name)

        for index, row in contradiction_df.iterrows():
            for column in categories:
                contradictions = json.loads(row[column])
                if len(contradictions):
                    temp_list = []
                    for contradict in contradictions:
                        question = get_question_with_number(contradict, self.test_questions)
                        value = original_df[original_df['İsim Soyisim'] == row['İsim Soyisim']][contradict].values[0]
                        text = text_df_not_concatted[text_df_not_concatted['İsim Soyisim'] == row['İsim Soyisim']][
                            question].values[0]
                        temp_list.append(f"Soru: {question} --- cevap: {value}: --- \"{text}\"")
                        row[column] = str(temp_list)
                else:
                    row[column] = ""

        for category in categories:
            contradiction_df.rename(columns={category: category + " contradict"}, inplace=True)

        contradiction_df.to_excel(f"{self.temp_dir}/{self.name}-contradict.xlsx", engine="openpyxl", index=False)

    def add_graphs_sheet(self, path=None):


        sheet_name = "graphs"
        path = path if path else self.output_path

        df = pd.read_excel(path, engine='openpyxl')

        # drop std and avg
        df.drop(df.tail(2).index,
                inplace=True)

        df = df[["İsim Soyisim"] + self.categories + self.categories_t]

        from openpyxl import load_workbook
        book = load_workbook(self.output_path)
        writer = pd.ExcelWriter(self.output_path, engine='openpyxl')
        writer.book = book
        df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
        writer.close()
        for i, category in enumerate(self.categories + self.categories_t):
            self.add_barchart(sheetname=sheet_name,
                              start_col=category,
                              end_col=category,
                              title=category,
                              name_col="İsim Soyisim",
                              max_row=len(df) + 1,
                              output_path=self.output_path,
                              gap=5,
                              multiplier=i+1)

    def filter_df_by_password_return(self, df):
        df = df[df['Şifre'].str.contains('|'.join(self.passwords))].copy()
        # standardize name column
        df['İsim Soyisim'] = df['Şifre'].apply(lambda x: self.passwords[x]).copy()
        return df

    def create_report(self):
        self.filter_df_by_password()
        if len(self.test_df) == 0:
            return 0
        # map answers to their corresponding score in dataframe
        self.map_answers()
        # remove questions from the result and add total column
        self.prepare_result()
        # add t-score
        self.add_tscore()
        # add trust score
        if self.should_be_same:  # If test has contradiction array
            self.add_trust_score()

        # add mean
        self.add_mean_row()
        # add standart deviation
        self.add_std_row()
        # round df
        self.round_df()
        # save
        self.save_report()
        # create 1 copy to creating colors and adjusting width
        self.copy_df = self.result_df.copy()
        # colorize
        self.format_report()
        # add category sheet
        self.add_graphs_sheet()
        # adjust width
        self.adjust_width()
        # create text report depending on answers
        if os.path.exists(self.text_file_path):
            self.create_text_report()

        print(f"Created report for: {self.name}")

    def mean_highlighter(self, x, y):
        very_good = f"background-color: {self.colors['very_good'] if self.colors['very_good'] else 'white'}; color: black;"
        good = f"background-color: {self.colors['good'] if self.colors['very_good'] else 'white'}; color: black;"
        okey = f"background-color: {self.colors['okey'] if self.colors['very_good'] else 'white'}; color: black;"
        bad = f"background-color: {self.colors['bad'] if self.colors['very_good'] else 'white'}; color: black;"
        very_bad = f"background-color: {self.colors['very_bad'] if self.colors['very_good'] else 'white'}; color: black;"

        if x > y * 1.25:
            return very_good
        if y * 1.25 > x >= y * 1.1:
            return good
        if y * 1.1 > x >= y * 0.85:
            return okey
        if y * 0.85 > x > y * 0.5:
            return bad
        if y * 0.5 > x:
            return very_bad
        else:
            return None

    def adjust_font(self, x):
        return f"font-size: {self.font}; font-weight: bold"

    def add_barchart(self, sheetname, start_col, end_col, name_col, output_path,
                     min_row=None,
                     max_row=None,
                     title=None,
                     total=None,
                     gap=None,
                     multiplier=1):

        df = pd.read_excel(output_path, engine='openpyxl', sheet_name=sheetname)
        min_col = df.columns.get_loc(start_col) + 1
        max_col = df.columns.get_loc(end_col) + 1

        min_row = 1 if not min_row else min_row
        max_row = len(df) - 1 if not max_row else max_row
        total = max_row if not total else total
        name_col = df.columns.get_loc(name_col) + 1

        data_len = max_row - min_row
        # bar chart adding
        wb = load_workbook(output_path)
        ws = wb[sheetname]

        if self.name == 'Performance':
            data = Reference(ws, min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row)
            titles = Reference(ws, min_col=name_col, max_col=name_col, min_row=min_row, max_row=max_row)
        else:
            data = Reference(ws, min_col=min_col, max_col=max_col, min_row=1, max_row=max_row)
            titles = Reference(ws, min_col=name_col, max_col=name_col, min_row=2, max_row=max_row)
        chart = BarChart3D()

        # title
        chart.title = self.name if not title else title
        man_layout = ManualLayout(xMode="edge", yMode="edge", x=0.0, y=0.0)
        chart.title.layout = Layout(manualLayout=man_layout)

        # data
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)

        # shape gap
        chart.gapWidth = gap if gap else 50
        chart.varyColors = "0000FFFF"
        chart.shape = 'cylinder'
        chart.bargroupgap = 20
        chart.height = 10  # default is 7.5
        chart.width = len(titles) * 3 if data_len < 15 else data_len * 2  # default is 15

        # datalabels
        chart.dataLabels = DataLabelList()
        chart.dataLabels.showVal = True

        # dataSeries
        chart.serLines = openpyxl.chart.axis.ChartLines()
        chart.serLines.showVal = True

        # text properteis
        axis = CharacterProperties(sz=1300, b=True)
        chart.serLines.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)])
        chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)])
        chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)])
        chart.dataLabels.textProperties = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis), endParaRPr=axis)])

        height_adjuster = (int(multiplier) - 1) * chart.height

        ws.add_chart(chart, f"B{total + height_adjuster * 2 + chart.height}")
        wb.save(output_path)
