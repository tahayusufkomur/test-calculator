import json
import pathlib
import string

import pandas as pd


class Attendance:
    name = "attendance"
    type = 2
    description = "Minesota"
    current_path = pathlib.Path(__file__).parent.resolve()
    passwords_path = f"{current_path}/../files/passwords.json"
    passwords = json.load(open(passwords_path, encoding="utf-8"))
    report = {}
    output_path = f"{current_path}/../raporlar/excel_reports/attendance_list.xlsx"
    font = '13pt;'
    colors = {'very_bad': "#FF00F4", 'bad': "#CD7F32", 'okey': "#C9CC3F", 'good': "#4CBB17",
              'very_good': "#40E0D0"}
    keys = {}

    def __init__(self, files):
        self.files = files
        self.EXTENDED_UPPERCASE_ALPHABET = list(string.ascii_uppercase) + ['A' + i for i in string.ascii_uppercase]

    def create_report(self):
        if "regex" in self.passwords.keys():
            del self.passwords['regex']
        # list to delete tests that no need attendance check
        del_list = []
        for file, path in self.files.items():
            try:
                self.keys[file] = pd.read_excel(path, engine='openpyxl')['Şifre'].tolist()
            except:
                print(f"{path} is empty")
                del_list.append(file)

        for file in del_list:
            del self.files[file]

        for password in self.passwords:
            row = {}
            for file, path in self.files.items():
                keys = self.keys[file]
                if password in keys:
                    row[file] = "Girdi"
                else:
                    row[file] = "Girmedi"

            self.report[self.passwords[password]] = row

        df = pd.DataFrame.from_dict(self.report, orient='index')
        df.to_excel(self.output_path, engine='openpyxl')

        # highlight değerlendirmeyenler
        output_path = self.output_path
        from openpyxl import load_workbook
        wb = load_workbook(output_path)
        # page 1
        sheet_name = "Sheet1"
        df = pd.read_excel(output_path, engine='openpyxl', sheet_name=sheet_name)
        writer = pd.ExcelWriter(output_path, engine='openpyxl', mode='a')
        df = df.style \
            .applymap(self.adjust_font) \
            .applymap(self.negative_highlighter)

        writer.book = wb
        wb.remove(wb[sheet_name])
        df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
        writer.close()

        self.adjust_width(output_path)

    def negative_highlighter(self, x):
        negative = f"background-color: {self.colors['very_bad'] if self.colors['very_good'] else 'white'}; color: black;"
        positive = f"background-color: {self.colors['very_good'] if self.colors['very_good'] else 'white'}; color: black;"
        if x == "Girdi":
            return positive
        if x == "Girmedi":
            return negative
        else:
            return None

    def adjust_width(self, path=None):
        from openpyxl import load_workbook
        if path:
            path = path
        else:
            path = self.output_path

        wb = load_workbook(path)
        for sheet_name in wb.sheetnames:
            sheet = wb.get_sheet_by_name(sheet_name)

            df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")

            excel_keymap = {}
            for i, column in enumerate(df.columns):
                excel_keymap[column] = self.EXTENDED_UPPERCASE_ALPHABET[i]
            # Auto-adjust columns' width
            for column in df.columns:
                column_width = max(df[column].astype(str).map(len).max(), len(column))
                sheet.column_dimensions[excel_keymap[column]].width = column_width + 3
        wb.save(path)

    def adjust_font(self, x):
        return f"font-size: {self.font}; font-weight: bold"


if __name__ == '__main__':
    files = {'B5KT': '/Users/tkoemue/ws/projects/merve/five-factor-test-calculator/yanıtlar/B5KT.xlsx',
             'PLO': '/Users/tkoemue/ws/projects/merve/five-factor-test-calculator/yanıtlar/PLO.xlsx'}

    attendance = Attendance(files=files)
    attendance.create_report()
