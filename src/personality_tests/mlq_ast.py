import json
import pathlib
import string
import numpy as np
import pandas as pd
from openpyxl import load_workbook

from src.base_personality_test import BasePersonalityTest
from src.utilities import get_number_from_question, is_question


class MlqAst(BasePersonalityTest):
    mlq_path = pathlib.Path(__file__).parent.resolve()

    def __init__(self, non_manager_path: str, manager_path: str, colors, passwords_path: str=None):
        self.passwords = json.load(open(passwords_path, encoding="utf-8")) if passwords_path else json.load(open(f"{self.mlq_path}/../files/passwords.json", encoding="utf-8"))
        self.exact_copy = None
        if colors is None:
            colors = {'very_bad': "#FF00F4", 'bad': "#CD7F32", 'okey': "#C9CC3F", 'good': "#4CBB17",
                      'very_good': "#40E0D0"}
        self.non_manager_path = non_manager_path
        self.manager_path = manager_path
        self.fix_excel_names(non_manager_path,
                             "Değerlendirdiğiniz kişinin adını-soyadını ve  ünvanını eksiksiz  olacak şekilde yazınız. ")
        self.fix_excel_names(manager_path,
                             "İsim -Soyisim ve ünvanınızı eksiksiz olarak yazınız. ")

        self.EXTENDED_UPPERCASE_ALPHABET = list(string.ascii_uppercase) + ['A' + i for i in string.ascii_uppercase]
        self.colors = colors
        self.managers = None
        self.test_df = self.read_test_answers()
        self.personal_questions, self.test_questions = self.parse_test()

    round = 1
    name = "mlq-ast"
    type = 2
    description = "MLQ Çoklu Liderlik Alt Faktörleri"

    scoring = {
        'Hiç bir zaman': 0,
        'Arada bİr': 1,
        'Bazen': 2,
        'Sık sık (genellikle)': 3,
        'Her zaman': 4
    }

    reverse_scoring = {
        'Hiç bir zaman': 4,
        'Arada bir': 3,
        'Bazen': 2,
        'Sık sık (genellikle)': 1,
        'Her zaman': 0
    }

    categories_dict = {1: ["İdeal Etki (atfedilen davranış)", [10, 18, 21, 25]],
                       2: ["İdeal Etki (davranıs)", [6, 14, 23, 34]],
                       3: ["İlham Verici Motivasyon", [9, 13, 26, 36]],
                       4: ["Entellektüel Uyarım / Zihinsel Teşvik", [2, 8, 30, 32]],
                       5: ["Bireysel Gelişime Destek", [15, 19, 29, 31]],
                       6: ["Koşulsal Ödül", [1, 11, 16, 35]],
                       7: ["İstisnalarla Yönetim (aktif)", [4, 22, 24, 27]],
                       8: ["İstisnalarla Yönetim (pasif)", [3, 12, 17, 20]],
                       9: ["Tam Serbesti Tanıyan Liderlik", [5, 7, 28, 33]],
                       10: ["Ekstra Çaba", [39, 42, 44]],
                       11: ["Etkinlik / Etkililik", [37, 40, 43]],
                       12: ["Tatmin (liderlik tarzı ile ilgili)", [38, 41, 45]]
                       }

    def read_test_answers(self):
        ast_df = pd.read_excel(self.non_manager_path, engine='openpyxl')
        ust_df = pd.read_excel(self.manager_path, engine='openpyxl')
        ust_df = self.filter_df_by_password_return(ust_df)

        ast_df.rename(columns={'Değerlendirdiğim kişiden daha alt konumdayım': 'Konum',
                               'Şifreyi giriniz ': 'Şifre',
                               'Değerlendirdiğiniz kişinin adını-soyadını ve  ünvanını eksiksiz  olacak şekilde yazınız. ': 'İsim Soyisim'}, inplace=True)
        ust_df.rename(columns={'Kendimi değerlendiriyorum.': 'Konum',
                               'İsim -Soyisim ve ünvanınızı eksiksiz olarak yazınız. ': 'İsim Soyisim'}, inplace=True)

        ast_df['Konum'] = ast_df['Konum'].map({'EVET': 'Ast'})
        ust_df['Konum'] = ust_df['Konum'].map({'EVET': 'Üst'})

        ust_df.columns = ast_df.columns
        df = ast_df.append(ust_df, ignore_index=True)
        self.managers = df[df['Konum'] == 'Üst']['İsim Soyisim'].tolist()
        return df

    def parse_test(self):
        columns = self.test_df.columns
        personal_questions = []
        test_questions = []

        for column in columns:
            if is_question(column):
                test_questions.append(column)
            else:
                personal_questions.append(column)
        return personal_questions, test_questions

    def filter_df_by_password_return(self, df):
        df = df[df['Şifre'].str.contains('|'.join(self.passwords))].copy()
        # standardize name column
        df['Şifre'] = df['Şifre'].apply(lambda x: self.passwords[x] if x in self.passwords else x).copy()
        return df

    def create_report(self):

        total_df = self.test_df.copy()
        total_df['Şifre'] = total_df['Şifre'].apply(lambda x: self.passwords[x] if x in self.passwords else x)
        for manager in self.managers:

            self.test_df = total_df[total_df['İsim Soyisim'] == manager].copy()
            self.output_path = f"{self.current_path}/raporlar/excel_reports/mlq_reports/{manager}-{self.name}-rapor-{self.date}.xlsx"

            # if there is less then 3 employee evaluated manager then skip
            if len(self.test_df) < 2:
                continue
            # # # map answers to their corresponding score in dataframe
            self.map_answers()
            # # remove questions from the result and add total column
            self.prepare_result()
            # add z-score
            self.add_tscore()
            # add mean
            self.add_mean_row()
            # add standart deviation
            self.add_std_row()
            # round df
            self.round_df()
            # save report with 1 page
            self.save_report()
            # add second page
            self.add_question_sheet()
            # add third page
            self.add_manager_report()
            # take a copy of result df
            self.copy_df = self.test_df.copy()
            # colorize
            self.format_report()
            # adjust width
            self.adjust_width()

    # only include asts for calculation
    def add_mean_row(self):
        self.result_df.loc['ortalama'] = self.test_df[self.test_df['Konum'] == 'Ast'].mean(numeric_only=True)

    # only include asts for calculation
    def add_std_row(self):
        self.result_df.loc['standart_sapma'] = self.test_df[self.test_df['Konum'] == 'Ast'].std(numeric_only=True)

    def map_answers(self):
        for column in self.test_questions:
            if column in [3, 12, 17, 20, 5, 7, 28, 33]:
                self.test_df[column] = self.test_df[column].map(self.reverse_scoring)
            else:
                self.test_df[column] = self.test_df[column].map(self.scoring)
        self.exact_copy = self.test_df.copy()

    def filter_df_by_password(self):
        self.test_df = self.test_df[self.test_df['Şifre'].str.contains('|'.join(self.passwords))]
        self.test_df['Şifre'] = self.test_df['Şifre'].map(self.passwords)

    def add_question_sheet(self):

        # TODO: use dictionary for building dataframe
        # an array to build dataframe
        summary_array = []

        df = self.exact_copy[
            self.exact_copy
            ['Konum'] == "Ast"
            ].copy()

        for key, category_row in self.categories_dict.items():
            for i in category_row[1]:
                category_name = category_row[0]
                question = self.test_questions[i-1]
                question_mean = round(df[question].mean(), self.round)
                category_mean = self.test_df[category_name].mean()
                question_mean = question_mean

                if key > 9:
                    category_mean = (category_mean/10).round(self.round)
                else:
                    category_mean = (category_mean/10).round(self.round)
                diff = abs(category_mean-question_mean)
                summary_array.append([category_name,
                                      question,
                                      question_mean,
                                      category_mean,
                                      diff])

        book = load_workbook(self.output_path)
        writer = pd.ExcelWriter(self.output_path, engine='openpyxl')
        writer.book = book

        columns = ["Kategori", "Soru", "Ortalama Puan", "Kategori Puanı", "Fark"]
        np_array = np.array(summary_array)

        df = pd.DataFrame(np_array, columns=columns, index=None)
        df = df.sort_values("Kategori")
        df = df.round(self.round)

        df.to_excel(excel_writer=writer, sheet_name="Soru Detayları", index=False)
        writer.close()

    def format_report(self):
        from openpyxl import load_workbook
        wb = load_workbook(self.output_path)

        # page 1
        sheet_name = "Sheet1"
        df = pd.read_excel(self.output_path, engine='openpyxl', sheet_name=sheet_name)
        writer = pd.ExcelWriter(self.output_path, engine='openpyxl', mode='a')
        column_1 = 'İdeal Etki (atfedilen davranış)'
        column_2 = 'İdeal Etki (davranıs)'
        column_3 = 'İlham Verici Motivasyon'
        column_4 = 'Entellektüel Uyarım / Zihinsel Teşvik'
        column_5 = 'Bireysel Gelişime Destek'
        column_6 = 'Koşulsal Ödül'
        column_7 = 'İstisnalarla Yönetim (aktif)'
        column_8 = 'İstisnalarla Yönetim (pasif)'
        column_9 = 'Tam Serbesti Tanıyan Liderlik'
        column_10 = 'Ekstra Çaba'
        column_11 = 'Etkinlik / Etkililik'
        column_12 = 'Tatmin (liderlik tarzı ile ilgili)'

        mean_1 = df[column_1].mean()
        mean_2 = df[column_2].mean()
        mean_3 = df[column_3].mean()
        mean_4 = df[column_4].mean()
        mean_5 = df[column_5].mean()
        mean_6 = df[column_6].mean()
        mean_7 = df[column_7].mean()
        mean_8 = df[column_8].mean()
        mean_9 = df[column_9].mean()
        mean_10 = df[column_10].mean()
        mean_11 = df[column_11].mean()
        mean_12 = df[column_12].mean()

        mean_t_1 = df[column_1 + " T"].mean()
        mean_t_2 = df[column_2 + " T"].mean()
        mean_t_3 = df[column_3 + " T"].mean()
        mean_t_4 = df[column_1 + " T"].mean()
        mean_t_5 = df[column_2 + " T"].mean()
        mean_t_6 = df[column_3 + " T"].mean()
        mean_t_7 = df[column_1 + " T"].mean()
        mean_t_8 = df[column_2 + " T"].mean()
        mean_t_9 = df[column_3 + " T"].mean()
        mean_t_10 = df[column_1 + " T"].mean()
        mean_t_11 = df[column_2 + " T"].mean()
        mean_t_12 = df[column_3 + " T"].mean()

        df = df.style \
            .applymap(self.adjust_font) \
            .applymap(lambda x: self.mean_highlighter(x, mean_1), subset=pd.IndexSlice[:, [column_1]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_2), subset=pd.IndexSlice[:, [column_2]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_3), subset=pd.IndexSlice[:, [column_3]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_4), subset=pd.IndexSlice[:, [column_4]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_5), subset=pd.IndexSlice[:, [column_5]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_6), subset=pd.IndexSlice[:, [column_6]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_7), subset=pd.IndexSlice[:, [column_7]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_8), subset=pd.IndexSlice[:, [column_8]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_9), subset=pd.IndexSlice[:, [column_9]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_10), subset=pd.IndexSlice[:, [column_10]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_11), subset=pd.IndexSlice[:, [column_11]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_12), subset=pd.IndexSlice[:, [column_12]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_t_1), subset=pd.IndexSlice[:, [column_1 + " T"]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_t_2), subset=pd.IndexSlice[:, [column_2 + " T"]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_t_3), subset=pd.IndexSlice[:, [column_3 + " T"]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_t_4), subset=pd.IndexSlice[:, [column_4 + " T"]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_t_5), subset=pd.IndexSlice[:, [column_5 + " T"]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_t_6), subset=pd.IndexSlice[:, [column_6 + " T"]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_t_7), subset=pd.IndexSlice[:, [column_7 + " T"]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_t_8), subset=pd.IndexSlice[:, [column_8 + " T"]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_t_9), subset=pd.IndexSlice[:, [column_9 + " T"]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_t_10), subset=pd.IndexSlice[:, [column_10 + " T"]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_t_11), subset=pd.IndexSlice[:, [column_11 + " T"]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_t_12), subset=pd.IndexSlice[:, [column_12 + " T"]])

        writer.book = wb
        wb.remove(wb[sheet_name])
        df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
        writer.close()

        # page 2
        sheet_name = "Soru Detayları"
        df = pd.read_excel(self.output_path, engine='openpyxl', sheet_name=sheet_name)
        writer = pd.ExcelWriter(self.output_path, engine='openpyxl', mode='a')

        column_1 = 'Ortalama Puan'
        column_2 = 'Kategori Puanı'
        column_3 = 'Fark'

        mean_1 = df[column_1].mean()
        mean_2 = df[column_2].mean()
        mean_3 = df[column_3].mean()

        df = df.style \
            .applymap(self.adjust_font) \
            .applymap(lambda x: self.mean_highlighter(x, mean_1), subset=pd.IndexSlice[:, [column_1]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_2), subset=pd.IndexSlice[:, [column_2]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_3), subset=pd.IndexSlice[:, [column_3]])

        writer.book = wb
        wb.remove(wb[sheet_name])
        df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
        writer.close()

        # page 3
        sheet_name = "Çalışanların Gözünden Yönetici"
        df = pd.read_excel(self.output_path, engine='openpyxl', sheet_name=sheet_name)
        writer = pd.ExcelWriter(self.output_path, engine='openpyxl', mode='a')

        column_1 = 'Çalışanların Değerlendirmesi'
        column_2 = 'Kendi Değerlendirmesi'
        column_3 = 'Fark'

        mean_1 = df[column_1].mean()
        mean_2 = df[column_2].mean()
        mean_3 = df[column_3].mean()

        df = df.style \
            .applymap(self.adjust_font) \
            .applymap(lambda x: self.mean_highlighter(x, mean_1), subset=pd.IndexSlice[:, [column_1]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_2), subset=pd.IndexSlice[:, [column_2]]) \
            .applymap(lambda x: self.mean_highlighter(x, mean_3), subset=pd.IndexSlice[:, [column_3]])

        writer.book = wb
        wb.remove(wb[sheet_name])
        df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
        writer.close()

        # add bar charts
        sheet_name = "Sheet1"
        df = pd.read_excel(self.output_path, engine='openpyxl', sheet_name=sheet_name)
        self.add_barchart(sheetname="Sheet1",
                          start_col='İdeal Etki (atfedilen davranış) T',
                          end_col='Tatmin (liderlik tarzı ile ilgili) T',
                          name_col='Şifre',
                          output_path=self.output_path,
                          max_row=len(df))

        sheet_name = "Soru Detayları"
        df = pd.read_excel(self.output_path, engine='openpyxl', sheet_name=sheet_name)
        self.add_barchart(sheetname="Soru Detayları",
                          start_col="Ortalama Puan",
                          end_col="Fark",
                          name_col='Soru',
                          output_path=self.output_path,
                          max_row=len(df))

        sheet_name = "Çalışanların Gözünden Yönetici"
        df = pd.read_excel(self.output_path, engine='openpyxl', sheet_name=sheet_name)
        self.add_barchart(sheetname="Çalışanların Gözünden Yönetici",
                          start_col='Çalışanların Değerlendirmesi',
                          end_col='Fark',
                          name_col='Kategoriler',
                          output_path=self.output_path,
                          max_row=len(df))

    def add_manager_report(self):
        sheet_name = "Çalışanların Gözünden Yönetici"

        df = self.exact_copy.copy()

        calisanlarin_degerlendirmesi = df[
            df[
                'Konum'] == "Ast"]

        kendi_degerlendirmesi = df[
            df['Konum'] == "Üst"]

        calisanlar = self.create_category_score(df=calisanlarin_degerlendirmesi)
        kendisi = self.create_category_score(df=kendi_degerlendirmesi)

        diff = []
        for i in range(len(calisanlar)):
            diff.append(abs(calisanlar[i][1]-kendisi[i][1]))

        report = {'Kategoriler': [x[0] for x in calisanlar],
                  'Çalışanların Değerlendirmesi': [x[1] for x in calisanlar],
                  'Kendi Değerlendirmesi': [x[1] for x in kendisi],
                  'Fark': diff}

        report_df = pd.DataFrame(report)

        book = load_workbook(self.output_path)
        writer = pd.ExcelWriter(self.output_path, engine='openpyxl')
        writer.book = book

        report_df = report_df.round(self.round)
        report_df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
        writer.close()

    def create_category_score(self, df):
        category_scores = []
        for key, row in self.categories_dict.items():
            category_score = 0
            for column in self.test_questions:
                number = get_number_from_question(column)
                if int(number) in row[1]:
                    category_score += df[column].mean()

            if key > 9:
                category_scores.append([row[0], category_score / 3])
            else:
                category_scores.append([row[0], category_score / 4])

        return category_scores

    def negative_highlighter(self, x):
        return f"background-color: {self.colors['very_bad'] if self.colors['very_good'] else 'white'}; color: black;"


if __name__ == '__main__':
    x = MlqAst('/Users/tkoemue/ws/projects/merve/five-factor-test-calculator/yanıtlar/MLQ-AST.xlsx',
               '/Users/tkoemue/ws/projects/merve/five-factor-test-calculator/yanıtlar/MLQ-UST.xlsx',
               None)

    x.create_report()